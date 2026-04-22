#!/usr/bin/env python3
"""
https://archive.ics.uci.edu/dataset/26/connect+4

Convert UCI Connect-4 positions (connect-4.data) into the row format used by
this project, and label each position with our Connect4Lookahead policy.

Chunked resume behavior
-----------------------
- Processes only CHUNK_SIZE source rows per run
- Appends exported rows into the same Excel workbook
- Stores resume state in sheet "_progress"
- Stores per-run history in sheet "_chunks"
- On rerun, continues from the next source row
- When finished, marks the workbook as finished

Input
-----
A local file named connect-4.data with rows like:
    b,b,b,...,x,o,...,win

The UCI dataset encodes 42 board cells in this order:
    a1,a2,...,a6,b1,...,g6,class
where row 1 is the bottom row and row 6 is the top row.

Output
------
One row per input position, in the same row schema used by your generator:
    label, reward, game, ply, action, p0..p6, player, 0-0..5-6
plus a few useful extras:
    uci_class, value_first_player, x_count, o_count

Board storage matches your existing pipeline:
- shape (6,7)
- top row is row 0
- values are {0,1,2}
- flattened columns are named "r-c" via top-row-first order
"""

from __future__ import annotations

from pathlib import Path
import sys
import random
from typing import Any, Dict, Iterable, List, Optional
from datetime import datetime
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

# -----------------------------------------------------------------------------
# One-click Spyder config
# -----------------------------------------------------------------------------
DATA_PATH = Path("connect-4.data")
OUTPUT_PATH = Path("UCI_connect4_L13.xlsx")
SHEET_NAME = "Sheet1"

LOOKAHEAD_DEPTH = 13
LABEL = f"UCI_LA{LOOKAHEAD_DEPTH}"
ACTION_MODE = "center"   # "center", "left", "sample"
SEED = 666
STORE_POLICY_PROBS = True
STORE_UCI_VALUE = True

CHUNK_SIZE = 10000
PROGRESS_EVERY = 100

PROGRESS_SHEET = "_progress"
CHUNKS_SHEET = "_chunks"

RESET_OUTPUT = False
STRICT_CONFIG_MATCH = True

# -----------------------------------------------------------------------------
# Imports: local module, with project root added to sys.path
# -----------------------------------------------------------------------------
_THIS_FILE = Path(__file__).resolve()
_PROJECT_ROOT = _THIS_FILE.parent.parent   # .../Connect4

if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from C4.fast_connect4_lookahead import Connect4Lookahead

ROWS = 6
COLS = 7
CENTER_COL = 3
CENTER_ORDER = (3, 4, 2, 5, 1, 6, 0)
COL_LABELS = "abcdefg"
ROW_LABELS = (1, 2, 3, 4, 5, 6)   # UCI uses bottom=1 ... top=6

UCI_COLUMNS = [f"{c}{r}" for c in COL_LABELS for r in ROW_LABELS]
TOKEN_MAP = {"b": 0, "x": 1, "o": 2}
CLASS_TO_VALUE = {"win": 1.0, "loss": -1.0, "draw": 0.0}


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _now_str() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _board_to_cells(board: np.ndarray) -> Dict[str, int]:
    """Flatten board into columns 0-0 .. 5-6, top row first."""
    out: Dict[str, int] = {}
    for r in range(ROWS):
        for c in range(COLS):
            out[f"{r}-{c}"] = int(board[r, c])
    return out


def _records_to_dataframe(records: List[Dict[str, Any]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()

    df = pd.DataFrame.from_records(records)

    board_cols = [f"{r}-{c}" for r in range(ROWS) for c in range(COLS)]
    preferred_order = (
        ["label", "reward", "game", "ply", "action"]
        + [f"p{a}" for a in range(COLS)]
        + ["player", "uci_class", "value_first_player", "x_count", "o_count"]
        + board_cols
    )

    existing = list(df.columns)
    ordered = [c for c in preferred_order if c in existing]
    tail = [c for c in existing if c not in ordered]

    return df[ordered + tail]


def _load_uci_positions(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path.resolve()}")

    cols = UCI_COLUMNS + ["class"]
    df = pd.read_csv(path, header=None, names=cols)
    return df


def _uci_row_to_board(row: pd.Series) -> np.ndarray:
    """
    Convert UCI order a1..g6 into project board order (6x7, top row first).

    UCI:
      a1 = bottom-left
      a6 = top-left
      g1 = bottom-right
      g6 = top-right

    Project board:
      board[0, 0] = top-left
      board[5, 6] = bottom-right
    """
    board = np.zeros((ROWS, COLS), dtype=np.int8)

    for c_idx, c in enumerate(COL_LABELS):
        for r_num in ROW_LABELS:
            key = f"{c}{r_num}"
            tok = str(row[key]).strip().lower()
            if tok not in TOKEN_MAP:
                raise ValueError(f"Unexpected token {tok!r} in column {key}")
            v = TOKEN_MAP[tok]
            project_r = ROWS - r_num
            board[project_r, c_idx] = v

    return board


def _infer_player_to_move(board: np.ndarray) -> int:
    """
    Infer side to move from counts, with project convention:
      1 = first player (x)
      2 = second player (o)

    Legal Connect-4 positions must satisfy:
      x_count == o_count      -> x to move
      x_count == o_count + 1  -> o to move
    """
    x_count = int(np.count_nonzero(board == 1))
    o_count = int(np.count_nonzero(board == 2))

    if x_count == o_count:
        return 1
    if x_count == o_count + 1:
        return 2

    raise ValueError(
        f"Illegal position counts: x_count={x_count}, o_count={o_count}"
    )


def _legal_cols(board: np.ndarray) -> List[int]:
    return [c for c in range(COLS) if int(board[0, c]) == 0]


def _uniform_among_best(scores: np.ndarray, legal_cols: Iterable[int]) -> np.ndarray:
    probs = np.zeros(COLS, dtype=np.float64)
    legal_cols = list(legal_cols)
    if not legal_cols:
        return probs

    legal_scores = np.array([scores[c] for c in legal_cols], dtype=np.float64)
    best = np.nanmax(legal_scores)

    best_cols = [
        c for c in legal_cols
        if np.isfinite(scores[c]) and np.isclose(scores[c], best, atol=1e-9, rtol=1e-9)
    ]
    if not best_cols:
        return probs

    p = 1.0 / float(len(best_cols))
    for c in best_cols:
        probs[c] = p
    return probs


def _choose_action_from_probs(probs: np.ndarray, mode: str, rng: random.Random) -> int:
    best_cols = [c for c in range(COLS) if float(probs[c]) > 0.0]
    if not best_cols:
        return CENTER_COL

    mode = str(mode).lower()

    if mode == "sample":
        total = float(probs.sum())
        x = rng.random() * total
        acc = 0.0
        for c in range(COLS):
            acc += float(probs[c])
            if x <= acc:
                return int(c)
        return int(best_cols[-1])

    if mode == "left":
        return int(min(best_cols))

    for c in CENTER_ORDER:
        if c in best_cols:
            return int(c)
    return int(best_cols[0])


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return bool(v)
    return str(v).strip().lower() in {"1", "true", "yes", "y"}


def _row_values(ws, row_idx: int) -> List[Any]:
    rows = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
    if not rows:
        return []
    return list(rows[0])


def _first_nonempty_row_idx(ws) -> Optional[int]:
    max_row = int(ws.max_row or 0)
    for r in range(1, max_row + 1):
        vals = _row_values(ws, r)
        if any(v is not None for v in vals):
            return r
    return None


def _normalize_leading_empty_rows(ws) -> None:
    """
    If a sheet starts with one or more empty rows, delete them so the first
    non-empty row becomes row 1.
    """
    first = _first_nonempty_row_idx(ws)
    if first is None:
        return
    if first > 1:
        ws.delete_rows(1, first - 1)


def _write_row(ws, row_idx: int, values: List[Any]) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


def _sheet_is_effectively_empty(ws) -> bool:
    return _first_nonempty_row_idx(ws) is None


def _read_progress_sheet(path: Path) -> Optional[Dict[str, Any]]:
    if not path.exists():
        return None

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if PROGRESS_SHEET not in wb.sheetnames:
            return None

        ws = wb[PROGRESS_SHEET]

        header_row = _first_nonempty_row_idx(ws)
        if header_row is None:
            return None

        value_row = header_row + 1
        if value_row > ws.max_row:
            return None

        headers = _row_values(ws, header_row)
        values = _row_values(ws, value_row)

        last_header_idx = -1
        for i, h in enumerate(headers):
            if h is not None:
                last_header_idx = i

        if last_header_idx < 0:
            return None

        headers = headers[:last_header_idx + 1]
        values = values[:last_header_idx + 1]

        progress = dict(zip(headers, values))

        if "finished" in progress:
            progress["finished"] = _to_bool(progress["finished"])

        return progress
    finally:
        wb.close()


def _validate_resume_config(
    progress: Dict[str, Any],
    *,
    data_path: Path,
    label: str,
    lookahead_depth: int,
) -> None:
    if not STRICT_CONFIG_MATCH:
        return

    old_data_path = str(progress.get("data_path", ""))
    old_label = str(progress.get("label", ""))
    old_depth = progress.get("lookahead_depth", None)

    if old_data_path and old_data_path != str(data_path.resolve()):
        raise ValueError(
            "Existing workbook was created with a different DATA_PATH.\n"
            f"Old: {old_data_path}\n"
            f"New: {data_path.resolve()}"
        )

    if old_label and old_label != str(label):
        raise ValueError(
            "Existing workbook was created with a different LABEL.\n"
            f"Old: {old_label}\n"
            f"New: {label}"
        )

    if old_depth is not None and int(old_depth) != int(lookahead_depth):
        raise ValueError(
            "Existing workbook was created with a different LOOKAHEAD_DEPTH.\n"
            f"Old: {old_depth}\n"
            f"New: {lookahead_depth}"
        )


def _build_progress_row(
    *,
    data_path: Path,
    output_path: Path,
    label: str,
    lookahead_depth: int,
    action_mode: str,
    seed: int,
    store_policy_probs: bool,
    store_uci_value: bool,
    total_input_rows: int,
    chunk_size: int,
    next_input_row: int,
    last_chunk_start: int,
    last_chunk_end_exclusive: int,
    last_chunk_exported_rows: int,
    last_chunk_skipped_rows: int,
    total_exported_rows: int,
    total_skipped_rows: int,
    finished: bool,
    created_at: Optional[str],
) -> Dict[str, Any]:
    return {
        "created_at": created_at or _now_str(),
        "updated_at": _now_str(),
        "data_path": str(data_path.resolve()),
        "output_path": str(output_path.resolve()),
        "sheet_name": SHEET_NAME,
        "label": str(label),
        "lookahead_depth": int(lookahead_depth),
        "action_mode": str(action_mode),
        "seed": int(seed),
        "store_policy_probs": bool(store_policy_probs),
        "store_uci_value": bool(store_uci_value),
        "total_input_rows": int(total_input_rows),
        "chunk_size": int(chunk_size),
        "last_chunk_start": int(last_chunk_start),
        "last_chunk_end_exclusive": int(last_chunk_end_exclusive),
        "next_input_row": int(next_input_row),
        "last_chunk_exported_rows": int(last_chunk_exported_rows),
        "last_chunk_skipped_rows": int(last_chunk_skipped_rows),
        "total_exported_rows": int(total_exported_rows),
        "total_skipped_rows": int(total_skipped_rows),
        "finished": bool(finished),
    }


def _append_df_and_update_meta(
    *,
    df_chunk: pd.DataFrame,
    output_path: Path,
    data_sheet_name: str,
    progress_row: Dict[str, Any],
    chunk_log_row: Dict[str, Any],
) -> None:
    """
    Append chunk rows to data sheet and replace/update metadata sheets
    in the same workbook.

    This version:
    - never creates a blank first row before header
    - can also repair older workbooks with leading blank rows
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet" and wb.active["A1"].value is None:
            wb.remove(wb.active)

    # -----------------------------------------------------------------
    # Data sheet
    # -----------------------------------------------------------------
    if data_sheet_name in wb.sheetnames:
        ws_data = wb[data_sheet_name]
    else:
        ws_data = wb.create_sheet(data_sheet_name)

    _normalize_leading_empty_rows(ws_data)

    if _sheet_is_effectively_empty(ws_data):
        if not df_chunk.empty:
            headers = list(df_chunk.columns)
            _write_row(ws_data, 1, headers)

            next_row = 2
            for row in df_chunk.itertuples(index=False, name=None):
                _write_row(ws_data, next_row, list(row))
                next_row += 1
    else:
        existing_header = _row_values(ws_data, 1)
        while existing_header and existing_header[-1] is None:
            existing_header.pop()

        if not df_chunk.empty:
            new_header = list(df_chunk.columns)

            if new_header != existing_header:
                if set(new_header) == set(existing_header):
                    df_chunk = df_chunk[existing_header]
                else:
                    raise ValueError(
                        "Chunk columns do not match existing Excel header.\n"
                        f"Existing: {existing_header}\n"
                        f"New: {new_header}"
                    )

            next_row = ws_data.max_row + 1
            for row in df_chunk.itertuples(index=False, name=None):
                _write_row(ws_data, next_row, list(row))
                next_row += 1

    # -----------------------------------------------------------------
    # Progress sheet
    # -----------------------------------------------------------------
    if PROGRESS_SHEET in wb.sheetnames:
        del wb[PROGRESS_SHEET]
    ws_progress = wb.create_sheet(PROGRESS_SHEET)

    progress_headers = list(progress_row.keys())
    progress_values = [progress_row[k] for k in progress_headers]
    _write_row(ws_progress, 1, progress_headers)
    _write_row(ws_progress, 2, progress_values)

    # -----------------------------------------------------------------
    # Chunk history sheet
    # -----------------------------------------------------------------
    if CHUNKS_SHEET in wb.sheetnames:
        ws_chunks = wb[CHUNKS_SHEET]
    else:
        ws_chunks = wb.create_sheet(CHUNKS_SHEET)

    _normalize_leading_empty_rows(ws_chunks)

    chunk_headers = list(chunk_log_row.keys())
    chunk_values = [chunk_log_row[k] for k in chunk_headers]

    if _sheet_is_effectively_empty(ws_chunks):
        _write_row(ws_chunks, 1, chunk_headers)
        _write_row(ws_chunks, 2, chunk_values)
    else:
        existing_chunk_header = _row_values(ws_chunks, 1)
        while existing_chunk_header and existing_chunk_header[-1] is None:
            existing_chunk_header.pop()

        if existing_chunk_header != chunk_headers:
            raise ValueError(
                "Existing _chunks header does not match current chunk log schema.\n"
                f"Existing: {existing_chunk_header}\n"
                f"New: {chunk_headers}"
            )

        next_row = ws_chunks.max_row + 1
        _write_row(ws_chunks, next_row, chunk_values)

    wb.save(output_path)
    wb.close()


# -----------------------------------------------------------------------------
# Main conversion, one chunk per run
# -----------------------------------------------------------------------------
def convert_uci_to_move_rows_chunked(
    *,
    data_path: Path,
    output_path: Path,
    lookahead_depth: int = 7,
    label: str = "UCI_LA7",
    action_mode: str = "center",
    seed: int = 666,
    store_policy_probs: bool = True,
    store_uci_value: bool = True,
    chunk_size: int = 5000,
    progress_every: int = 500,
) -> Dict[str, Any]:
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("This chunked/resume version is designed to use .xlsx output.")

    if RESET_OUTPUT and output_path.exists():
        output_path.unlink()

    rng = random.Random(seed)
    la = Connect4Lookahead()
    src = _load_uci_positions(data_path)
    total_rows = int(len(src))

    progress = _read_progress_sheet(output_path)

    if progress is None:
        start_idx = 0
        total_exported_rows = 0
        total_skipped_rows = 0
        created_at = _now_str()
        finished = False
        print("No previous progress found, starting from source row 0")
    else:
        _validate_resume_config(
            progress,
            data_path=data_path,
            label=label,
            lookahead_depth=lookahead_depth,
        )

        finished = bool(progress.get("finished", False))
        if finished:
            return {
                "status": "already_finished",
                "input_rows": total_rows,
                "exported_rows_total": int(progress.get("total_exported_rows", 0)),
                "skipped_rows_total": int(progress.get("total_skipped_rows", 0)),
                "next_input_row": int(progress.get("next_input_row", total_rows)),
                "lookahead_depth": int(lookahead_depth),
                "output_path": str(output_path),
            }

        start_idx = int(progress.get("next_input_row", 0))
        total_exported_rows = int(progress.get("total_exported_rows", 0))
        total_skipped_rows = int(progress.get("total_skipped_rows", 0))
        created_at = progress.get("created_at", _now_str())

        print(f"Resuming from source row {start_idx:,}")

    if start_idx >= total_rows:
        finished = True

        progress_row = _build_progress_row(
            data_path=data_path,
            output_path=output_path,
            label=label,
            lookahead_depth=lookahead_depth,
            action_mode=action_mode,
            seed=seed,
            store_policy_probs=store_policy_probs,
            store_uci_value=store_uci_value,
            total_input_rows=total_rows,
            chunk_size=chunk_size,
            next_input_row=total_rows,
            last_chunk_start=start_idx,
            last_chunk_end_exclusive=start_idx,
            last_chunk_exported_rows=0,
            last_chunk_skipped_rows=0,
            total_exported_rows=total_exported_rows,
            total_skipped_rows=total_skipped_rows,
            finished=True,
            created_at=created_at,
        )

        chunk_log_row = {
            "timestamp": _now_str(),
            "label": label,
            "lookahead_depth": int(lookahead_depth),
            "start_input_row": int(start_idx),
            "end_input_row_exclusive": int(start_idx),
            "requested_input_rows": 0,
            "exported_rows": 0,
            "skipped_rows": 0,
            "finished_after_chunk": True,
            "note": "No work, already at end of source file.",
        }

        _append_df_and_update_meta(
            df_chunk=pd.DataFrame(),
            output_path=output_path,
            data_sheet_name=SHEET_NAME,
            progress_row=progress_row,
            chunk_log_row=chunk_log_row,
        )

        return {
            "status": "finished_now",
            "input_rows": total_rows,
            "exported_rows_total": total_exported_rows,
            "skipped_rows_total": total_skipped_rows,
            "chunk_start": start_idx,
            "chunk_end_exclusive": start_idx,
            "chunk_exported_rows": 0,
            "chunk_skipped_rows": 0,
            "lookahead_depth": int(lookahead_depth),
            "output_path": str(output_path),
        }

    end_idx = min(start_idx + int(chunk_size), total_rows)

    print(
        f"Processing source rows [{start_idx:,}, {end_idx:,}) "
        f"out of {total_rows:,} total, depth={lookahead_depth}"
    )

    records: List[Dict[str, Any]] = []
    chunk_skipped = 0

    for idx in range(start_idx, end_idx):
        row = src.iloc[idx]

        try:
            board = _uci_row_to_board(row)
            player = _infer_player_to_move(board)
            legal = _legal_cols(board)
            if not legal:
                chunk_skipped += 1
                continue

            scores = la.n_step_action_scores(board, player, depth=int(lookahead_depth))
            probs = _uniform_among_best(scores, legal)
            action = _choose_action_from_probs(probs, mode=action_mode, rng=rng)

            x_count = int(np.count_nonzero(board == 1))
            o_count = int(np.count_nonzero(board == 2))
            ply = x_count + o_count + 1

            rec: Dict[str, Any] = {
                "label": str(label),
                "reward": 0.0,
                "game": int(idx),
                "ply": int(ply),
                "action": int(action),
                "player": int(player),
                **_board_to_cells(board),
            }

            if store_policy_probs:
                for a in range(COLS):
                    rec[f"p{a}"] = float(probs[a])

            if store_uci_value:
                uci_class = str(row["class"]).strip().lower()
                rec["uci_class"] = uci_class
                rec["value_first_player"] = float(CLASS_TO_VALUE.get(uci_class, np.nan))
                rec["x_count"] = int(x_count)
                rec["o_count"] = int(o_count)

            records.append(rec)

            processed_local = idx - start_idx + 1
            if progress_every > 0 and (processed_local % int(progress_every) == 0):
                print(
                    f"  source {idx + 1:,}/{total_rows:,}, "
                    f"chunk processed {processed_local:,}/{end_idx - start_idx:,}, "
                    f"chunk exported {len(records):,}, chunk skipped {chunk_skipped:,}"
                )

        except Exception as e:
            chunk_skipped += 1
            print(f"Skipping row {idx}: {e}")

    df_chunk = _records_to_dataframe(records)

    total_exported_rows += int(len(df_chunk))
    total_skipped_rows += int(chunk_skipped)
    next_input_row = int(end_idx)
    finished = next_input_row >= total_rows

    progress_row = _build_progress_row(
        data_path=data_path,
        output_path=output_path,
        label=label,
        lookahead_depth=lookahead_depth,
        action_mode=action_mode,
        seed=seed,
        store_policy_probs=store_policy_probs,
        store_uci_value=store_uci_value,
        total_input_rows=total_rows,
        chunk_size=chunk_size,
        next_input_row=next_input_row,
        last_chunk_start=start_idx,
        last_chunk_end_exclusive=end_idx,
        last_chunk_exported_rows=len(df_chunk),
        last_chunk_skipped_rows=chunk_skipped,
        total_exported_rows=total_exported_rows,
        total_skipped_rows=total_skipped_rows,
        finished=finished,
        created_at=created_at,
    )

    chunk_log_row = {
        "timestamp": _now_str(),
        "label": label,
        "lookahead_depth": int(lookahead_depth),
        "start_input_row": int(start_idx),
        "end_input_row_exclusive": int(end_idx),
        "requested_input_rows": int(end_idx - start_idx),
        "exported_rows": int(len(df_chunk)),
        "skipped_rows": int(chunk_skipped),
        "finished_after_chunk": bool(finished),
        "note": "",
    }

    _append_df_and_update_meta(
        df_chunk=df_chunk,
        output_path=output_path,
        data_sheet_name=SHEET_NAME,
        progress_row=progress_row,
        chunk_log_row=chunk_log_row,
    )

    summary = {
        "status": "chunk_done",
        "input_rows": int(total_rows),
        "chunk_start": int(start_idx),
        "chunk_end_exclusive": int(end_idx),
        "chunk_requested_rows": int(end_idx - start_idx),
        "chunk_exported_rows": int(len(df_chunk)),
        "chunk_skipped_rows": int(chunk_skipped),
        "next_input_row": int(next_input_row),
        "exported_rows_total": int(total_exported_rows),
        "skipped_rows_total": int(total_skipped_rows),
        "finished": bool(finished),
        "lookahead_depth": int(lookahead_depth),
        "output_path": str(output_path),
    }
    return summary


if __name__ == "__main__":
    start_time = time.time()

    summary = convert_uci_to_move_rows_chunked(
        data_path=DATA_PATH,
        output_path=OUTPUT_PATH,
        lookahead_depth=LOOKAHEAD_DEPTH,
        label=LABEL,
        action_mode=ACTION_MODE,
        seed=SEED,
        store_policy_probs=STORE_POLICY_PROBS,
        store_uci_value=STORE_UCI_VALUE,
        chunk_size=CHUNK_SIZE,
        progress_every=PROGRESS_EVERY,
    )

    print("\nDone.")
    for k, v in summary.items():
        print(f"{k}: {v}")

    elapsed_min = (time.time() - start_time) / 60.0
    print(f"\nThis run completed in {elapsed_min:.1f} minutes")