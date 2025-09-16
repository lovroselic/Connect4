# a0_logger_xlsx.py
from __future__ import annotations
import os
from typing import Iterable, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class ExcelLogger:
    """
    Append-only Excel logger. Writes rows to a given sheet, creating the file/sheet and header if needed.
    Usage:
        logger = ExcelLogger("runs/logs.xlsx", sheet_name="train", header=[...])
        logger.log(row_dict, fields=[...])   # appends one row (ordered by 'fields')
        logger.close()
    """
    def __init__(self, path: str, sheet_name: str, header: Iterable[str]):
        self.path = path
        self.sheet_name = sheet_name
        self.header = list(header)

        os.makedirs(os.path.dirname(path), exist_ok=True)

        if os.path.exists(path):
            wb = load_workbook(path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # if sheet is empty, write header
                if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1,1).value is None:
                    ws.append(self.header)
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(self.header)
        else:
            wb = Workbook()
            # Replace default 'Sheet' with our sheet
            ws: Worksheet = wb.active
            ws.title = sheet_name
            ws.append(self.header)
        self._wb = wb
        self._ws = ws
        self._save()

    def log(self, row: Dict[str, Any], fields: Iterable[str]):
        ordered = [row.get(k, None) for k in fields]
        self._ws.append(ordered)
        self._save()

    def _save(self):
        self._wb.save(self.path)

    def close(self):
        try:
            self._wb.save(self.path)
        finally:
            self._wb.close()
